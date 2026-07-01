# -*- coding: utf-8 -*-
import os, json, hashlib
from datetime import datetime
from functools import wraps
from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, send_file, abort)
from pdf_gen import generi_pdf
import io

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'cmn-predracun-tajni-kljuc-2026')

from db import q

def sha(pw): return hashlib.sha256(pw.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin'):
            flash('Nemate pristup ovoj stranici.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        u = q("SELECT * FROM korisnici WHERE username=? AND aktivan=1",
              (request.form['username'],), one=True)
        if u and u['password_hash'] == sha(request.form['password']):
            session['user_id'] = u['id']
            session['user_ime'] = u['ime']
            session['admin'] = bool(u['admin'])
            return redirect(url_for('index'))
        flash('Pogrešno korisničko ime ili lozinka.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    return redirect(url_for('novi_predracun'))

@app.route('/predracun/novi')
@login_required
def novi_predracun():
    dogadjaji = q("SELECT * FROM dogadjaji WHERE status='AKTIVNA' ORDER BY id DESC")
    cenovnik  = q("SELECT * FROM cenovnik WHERE status='AKTIVNA' ORDER BY naziv")
    last = q("SELECT broj FROM evidencija ORDER BY id DESC LIMIT 1", one=True)
    next_br = _next_broj(last['broj'] if last else None)
    today = datetime.now().strftime('%d.%m.%Y.')
    return render_template('novi_predracun.html',
        dogadjaji=dogadjaji, cenovnik=cenovnik,
        next_br=next_br, today=today)

def _next_broj(last, prefiks='PS'):
    import re
    yr = datetime.now().year % 100
    if not last:
        if prefiks:
            return f"{prefiks}-1/{yr:02d}"
        return f"1/{yr:02d}"
    m = re.search(r'(\d+)', last)
    n = int(m.group(1)) + 1 if m else 1
    if prefiks:
        return f"{prefiks}-{n}/{yr:02d}"
    return f"{n}/{yr:02d}"

@app.route('/api/next_broj')
@login_required
def api_next_broj():
    prefiks = request.args.get('prefiks', 'PS')
    last = q("SELECT broj FROM evidencija ORDER BY id DESC LIMIT 1", one=True)
    next_br = _next_broj(last['broj'] if last else None, prefiks)
    return jsonify({'broj': next_br})

@app.route('/predracun/sacuvaj', methods=['POST'])
@login_required
def sacuvaj_predracun():
    data = request.json
    stavke_json = json.dumps(data.get('stavke', []), ensure_ascii=False)
    osnov = 0.0; pdv_t = 0.0
    for st in data.get('stavke', []):
        try:
            c=float(st['cena']); k=float(st['kolicina']); p=float(st['pdv'])
            b=c*k; osnov+=b; pdv_t+=b*p/100
        except: pass
    ukupno = osnov + pdv_t
    dog_naziv = ''; dog_datum = ''
    if data.get('dogadjaj_id'):
        dog = q("SELECT * FROM dogadjaji WHERE id=?", (data['dogadjaj_id'],), one=True)
        if dog:
            dog_naziv = dog['naziv']; dog_datum = dog['datum']
    # Specifikacija dolazi iz forme (vezana za cenovnik stavku)
    specifikacija = data.get('specifikacija','')
    import os as _os
    _db_url = _os.environ.get('DATABASE_URL','')
    _sql = """INSERT INTO evidencija
        (broj, datum_izd, datum_val, kupac_naziv, kupac_pib, kupac_mb,
         kupac_adresa, kupac_mesto, kupac_jbkjs, kupac_extra,
         dogadjaj_id, dogadjaj_naziv, osnov, pdv, ukupno,
         status, napomena, specifikacija, stavke, korisnik_id, korisnik_ime, napomena_predracun)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
    if _db_url:
        _sql += " RETURNING id"
    lid = q(_sql,
        (data['broj'], data['datum_izd'], data['datum_val'],
         data['kupac_naziv'], data.get('kupac_pib',''), data.get('kupac_mb',''),
         data.get('kupac_adresa',''), data.get('kupac_mesto',''), data.get('kupac_jbkjs',''),
         data.get('kupac_extra',''),
         data.get('dogadjaj_id') or None, dog_naziv,
         osnov, pdv_t, ukupno,
         'ČEKA', '', specifikacija, stavke_json,
         session['user_id'], session['user_ime'],
         data.get('napomena','')), commit=True)
    return jsonify({'ok': True, 'id': lid})

@app.route('/predracun/pdf/<int:pid>')
@login_required
def predracun_pdf(pid):
    p = q("SELECT * FROM evidencija WHERE id=?", (pid,), one=True)
    if not p: abort(404)
    dog = q("SELECT datum FROM dogadjaji WHERE id=?", (p['dogadjaj_id'],), one=True) if p['dogadjaj_id'] else None
    data = dict(p); data['dogadjaj_datum'] = dog['datum'] if dog else ''
    pdf_bytes = generi_pdf(data)
    filename = f"{p['broj'].replace('/','-')}.pdf"
    return send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf',
                     as_attachment=False, download_name=filename)

@app.route('/predracun/pdf/download/<int:pid>')
@login_required
def predracun_pdf_download(pid):
    p = q("SELECT * FROM evidencija WHERE id=?", (pid,), one=True)
    if not p: abort(404)
    dog = q("SELECT datum FROM dogadjaji WHERE id=?", (p['dogadjaj_id'],), one=True) if p['dogadjaj_id'] else None
    data = dict(p); data['dogadjaj_datum'] = dog['datum'] if dog else ''
    pdf_bytes = generi_pdf(data)
    filename = f"{p['broj'].replace('/','-')}.pdf"
    return send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf',
                     as_attachment=True, download_name=filename)

@app.route('/evidencija')
@login_required
def evidencija():
    dog_filter = request.args.get('dogadjaj', '')
    status_filter = request.args.get('status', '')
    search = request.args.get('q', '')
    sql = "SELECT * FROM evidencija WHERE storno=0"
    args = []
    if dog_filter:
        sql += " AND dogadjaj_id=?"; args.append(dog_filter)
    if status_filter:
        sql += " AND status=?"; args.append(status_filter)
    if search:
        sql += " AND (kupac_naziv LIKE ? OR broj LIKE ? OR kupac_pib LIKE ?)"
        s = f'%{search}%'; args += [s, s, s]
    sql += " ORDER BY id DESC"
    rows = q(sql, args)
    dogadjaji = q("SELECT DISTINCT dogadjaj_id, dogadjaj_naziv FROM evidencija WHERE dogadjaj_id IS NOT NULL")

    def sums_for(status_check):
        matched = [r for r in rows if status_check(str(r['status'] or '').upper())]
        return {
            'count':  len(matched),
            'osnov':  sum(float(r['osnov'] or 0) for r in matched),
            'pdv':    sum(float(r['pdv']   or 0) for r in matched),
            'ukupno': sum(float(r['ukupno'] or 0) for r in matched),
        }

    stat_ukupno   = sums_for(lambda s: True)
    stat_aktivni  = sums_for(lambda s: s not in ('PLAĆENA','STORNO'))
    stat_placeni  = sums_for(lambda s: 'PLA' in s)
    stat_storno   = sums_for(lambda s: s == 'STORNO')

    return render_template('evidencija.html', rows=rows, dogadjaji=dogadjaji,
        dog_filter=dog_filter, status_filter=status_filter, search=search,
        stat_ukupno=stat_ukupno, stat_aktivni=stat_aktivni,
        stat_placeni=stat_placeni, stat_storno=stat_storno)

@app.route('/evidencija/status/<int:pid>', methods=['POST'])
@login_required
def set_status(pid):
    q("UPDATE evidencija SET status=? WHERE id=?", (request.json.get('status'), pid), commit=True)
    return jsonify({'ok': True})

@app.route('/evidencija/storno/<int:pid>', methods=['POST'])
@login_required
def storno(pid):
    q("UPDATE evidencija SET storno=1, status='STORNO' WHERE id=?", (pid,), commit=True)
    return jsonify({'ok': True})

@app.route('/evidencija/napomena/<int:pid>', methods=['POST'])
@login_required
def set_napomena(pid):
    q("UPDATE evidencija SET napomena=? WHERE id=?", (request.json.get('napomena',''), pid), commit=True)
    return jsonify({'ok': True})

@app.route('/kupci')
@login_required
def kupci():
    search = request.args.get('q', '').strip()
    if search:
        s = f'%{search}%'
        rows = q("SELECT * FROM kupci WHERE naziv LIKE ? OR pib LIKE ? OR jbkjs LIKE ? ORDER BY naziv LIMIT 100", (s, s, s))
    else:
        rows = q("SELECT * FROM kupci ORDER BY naziv LIMIT 100")
    total = q("SELECT COUNT(*) as c FROM kupci", one=True)['c']
    return render_template('kupci.html', rows=rows, search=search, total=total)

@app.route('/api/kupci/search')
@login_required
def api_kupci_search():
    term = request.args.get('q', '').strip()
    if not term: return jsonify([])
    s = f'%{term}%'
    rows = q("SELECT * FROM kupci WHERE naziv LIKE ? OR pib LIKE ? OR jbkjs LIKE ? ORDER BY naziv LIMIT 20", (s, s, s))
    return jsonify([dict(r) for r in rows])

@app.route('/kupci/novi', methods=['POST'])
@login_required
def novi_kupac():
    d = request.json
    q("INSERT INTO kupci (naziv, pib, mb, adresa, mesto, jbkjs) VALUES (?,?,?,?,?,?)",
      (d['naziv'], d.get('pib',''), d.get('mb',''), d.get('adresa',''), d.get('mesto',''), d.get('jbkjs','')), commit=True)
    return jsonify({'ok': True})

@app.route('/cenovnik')
@login_required
def cenovnik():
    prikazi_sve = request.args.get('sve', '0') == '1'
    rows = q("SELECT * FROM cenovnik ORDER BY naziv") if prikazi_sve else q("SELECT * FROM cenovnik WHERE status='AKTIVNA' ORDER BY naziv")
    return render_template('cenovnik.html', rows=rows, prikazi_sve=prikazi_sve)

@app.route('/cenovnik/novi', methods=['POST'])
@login_required
def nova_cena():
    d = request.json
    q("INSERT INTO cenovnik (naziv, cena, pdv, status) VALUES (?,?,?,?)",
      (d['naziv'], float(d['cena']), int(d.get('pdv',20)), 'AKTIVNA'), commit=True)
    return jsonify({'ok': True})

@app.route('/cenovnik/uredi/<int:cid>', methods=['POST'])
@login_required
def uredi_cenu(cid):
    d = request.json
    q("UPDATE cenovnik SET naziv=?, cena=?, pdv=?, specifikacija=? WHERE id=?",
      (d['naziv'], float(d['cena']), int(d.get('pdv',20)), d.get('specifikacija',''), cid), commit=True)
    return jsonify({'ok': True})

@app.route('/api/cenovnik/<int:cid>')
@login_required
def api_cenovnik_item(cid):
    r = q("SELECT * FROM cenovnik WHERE id=?", (cid,), one=True)
    if not r: return jsonify({})
    return jsonify(dict(r))

@app.route('/cenovnik/status/<int:cid>', methods=['POST'])
@login_required
def set_cena_status(cid):
    q("UPDATE cenovnik SET status=? WHERE id=?", (request.json['status'], cid), commit=True)
    return jsonify({'ok': True})

@app.route('/dogadjaji')
@login_required
def dogadjaji():
    prikazi_sve = request.args.get('sve', '0') == '1'
    rows = q("SELECT * FROM dogadjaji ORDER BY id DESC")
    if not prikazi_sve:
        rows = [r for r in rows if r['status'] == 'AKTIVNA']
    return render_template('dogadjaji.html', rows=rows, prikazi_sve=prikazi_sve)

@app.route('/dogadjaji/novi', methods=['POST'])
@login_required
def novi_dogadjaj():
    d = request.json
    q("INSERT INTO dogadjaji (naziv, datum, specifikacija, status) VALUES (?,?,?,?)",
      (d['naziv'], d.get('datum',''), d.get('specifikacija',''), 'AKTIVNA'), commit=True)
    return jsonify({'ok': True})

@app.route('/dogadjaji/status/<int:did>', methods=['POST'])
@login_required
def set_dog_status(did):
    q("UPDATE dogadjaji SET status=? WHERE id=?", (request.json['status'], did), commit=True)
    return jsonify({'ok': True})

@app.route('/backup/db')
@login_required
@admin_required
def backup_db():
    """Download backup as Excel file."""
    import pandas as pd
    from datetime import datetime
    DATABASE_URL = os.environ.get('DATABASE_URL', '')

    try:
        # Export all tables to Excel
        evidencija_rows = q("SELECT id,broj,datum_izd,datum_val,kupac_naziv,kupac_pib,kupac_mb,kupac_adresa,kupac_mesto,kupac_jbkjs,kupac_extra,dogadjaj_id,dogadjaj_naziv,osnov,pdv,ukupno,status,napomena,napomena_predracun,specifikacija,stavke,korisnik_ime,storno FROM evidencija ORDER BY id")
        kupci_rows     = q("SELECT naziv,pib,mb,adresa,mesto,jbkjs FROM kupci ORDER BY naziv")
        cenovnik_rows  = q("SELECT naziv,cena,pdv,status,specifikacija FROM cenovnik ORDER BY naziv")
        dogadjaji_rows = q("SELECT id,naziv,datum,status,specifikacija FROM dogadjaji ORDER BY id")

        def rows_to_df(rows):
            if not rows: return pd.DataFrame()
            return pd.DataFrame(rows)

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            rows_to_df(kupci_rows).to_excel(writer, sheet_name='Kupci', index=False)
            rows_to_df(cenovnik_rows).to_excel(writer, sheet_name='Cenovnik', index=False)
            rows_to_df(dogadjaji_rows).to_excel(writer, sheet_name='Dogadjaji', index=False)
            rows_to_df(evidencija_rows).to_excel(writer, sheet_name='Evidencija', index=False)
        buf.seek(0)

        fname = f"cmn_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=fname)
    except Exception as e:
        return f"Greška pri backup-u: {str(e)}", 500

@app.route('/korisnici')
@login_required
@admin_required
def korisnici():
    rows = q("SELECT * FROM korisnici ORDER BY ime")
    return render_template('korisnici.html', rows=rows)

@app.route('/korisnici/novi', methods=['POST'])
@login_required
@admin_required
def novi_korisnik():
    d = request.json
    try:
        q("INSERT INTO korisnici (ime, username, password_hash, admin) VALUES (?,?,?,?)",
          (d['ime'], d['username'], sha(d['password']), 1 if d.get('admin') else 0), commit=True)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/korisnici/reset/<int:uid>', methods=['POST'])
@login_required
@admin_required
def reset_lozinka(uid):
    q("UPDATE korisnici SET password_hash=? WHERE id=?", (sha(request.json.get('password','')), uid), commit=True)
    return jsonify({'ok': True})

@app.route('/korisnici/toggle/<int:uid>', methods=['POST'])
@login_required
@admin_required
def toggle_korisnik(uid):
    u = q("SELECT aktivan FROM korisnici WHERE id=?", (uid,), one=True)
    q("UPDATE korisnici SET aktivan=? WHERE id=?", (0 if u['aktivan'] else 1, uid), commit=True)
    return jsonify({'ok': True})

@app.route('/moja_lozinka', methods=['POST'])
@login_required
def moja_lozinka():
    d = request.json
    u = q("SELECT * FROM korisnici WHERE id=?", (session['user_id'],), one=True)
    if u['password_hash'] != sha(d['stara']):
        return jsonify({'ok': False, 'error': 'Pogrešna stara lozinka'})
    q("UPDATE korisnici SET password_hash=? WHERE id=?", (sha(d['nova']), session['user_id']), commit=True)
    return jsonify({'ok': True})

@app.template_filter('fmt_br')
def fmt_br_filter(x):
    try:
        return "{:,.2f}".format(float(x or 0)).replace(",","X").replace(".",",").replace("X",".")
    except:
        return "0,00"

@app.template_filter('fmt_datum')
def fmt_datum_filter(t):
    if not t: return ""
    s = str(t).strip()
    return s if s.endswith('.') else s


# ── EXPORT EVIDENCIJA ─────────────────────────────────────────────────────────
def _get_evidencija_rows(request):
    """Return filtered rows same as evidencija view."""
    dog_filter    = request.args.get('dogadjaj', '')
    status_filter = request.args.get('status', '')
    search        = request.args.get('q', '')
    sql  = "SELECT * FROM evidencija WHERE 1=1"
    args = []
    if dog_filter:
        sql += " AND dogadjaj_id=?"; args.append(dog_filter)
    if status_filter:
        sql += " AND status=?"; args.append(status_filter)
    if search:
        sql += " AND (kupac_naziv LIKE ? OR broj LIKE ? OR kupac_pib LIKE ?)"
        s = f'%{search}%'; args += [s, s, s]
    sql += " ORDER BY id DESC"
    return q(sql, args)

@app.route('/evidencija/export/xlsx')
@login_required
def export_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    rows = _get_evidencija_rows(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evidencija predracuna"

    thin = Side(style='thin', color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── SPECIFIKACIJA NA VRHU ────────────────────────────────────────────────
    from datetime import datetime as dt
    def sums_for_x(status_check):
        matched = [r for r in rows if status_check(str(r['status'] or '').upper())]
        return {'count': len(matched),
                'osnov':  sum(float(r['osnov'] or 0) for r in matched),
                'pdv':    sum(float(r['pdv']   or 0) for r in matched),
                'ukupno': sum(float(r['ukupno'] or 0) for r in matched)}

    st_svi     = sums_for_x(lambda s: True)
    st_aktiv   = sums_for_x(lambda s: s not in ('PLAĆENA','STORNO'))
    st_placeni = sums_for_x(lambda s: 'PLA' in s)
    st_storno  = sums_for_x(lambda s: s == 'STORNO')

    spec_title_font = Font(bold=True, size=13, color="1565C0")
    spec_hdr_font   = Font(bold=True, size=9, color="FFFFFF")
    spec_val_font   = Font(size=9)
    spec_num_font   = Font(size=9, bold=True)
    num_fmt = '#,##0.00'

    fills = {
        'svi':     PatternFill("solid", fgColor="1565C0"),
        'aktiv':   PatternFill("solid", fgColor="E65100"),
        'placeni': PatternFill("solid", fgColor="2E7D32"),
        'storno':  PatternFill("solid", fgColor="C62828"),
    }
    fonts_col = {
        'svi':     Font(bold=True, size=9, color="1A3A6B"),
        'aktiv':   Font(bold=True, size=9, color="7B2500"),
        'placeni': Font(bold=True, size=9, color="1B5E20"),
        'storno':  Font(bold=True, size=9, color="7F0000"),
    }
    bg = {
        'svi':     PatternFill("solid", fgColor="EEF3FB"),
        'aktiv':   PatternFill("solid", fgColor="FFF3E0"),
        'placeni': PatternFill("solid", fgColor="E8F5E9"),
        'storno':  PatternFill("solid", fgColor="FFEBEE"),
    }

    # Row 1: Title
    ws.merge_cells('A1:M1')
    c = ws['A1']
    c.value = f"EVIDENCIJA PREDRAČUNA  —  Izvoz: {dt.now().strftime('%d.%m.%Y. %H:%M')}"
    c.font  = spec_title_font
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    # Row 2: empty
    ws.row_dimensions[2].height = 6

    # Row 3: spec headers
    spec_cols = ['A','D','G','J']
    spec_keys = ['svi','aktiv','placeni','storno']
    spec_labels = ['SVI PREDRAČUNI','AKTIVNI (ČEKAJU)','PLAĆENI','STORNIRANI']
    spec_data = [st_svi, st_aktiv, st_placeni, st_storno]

    for sc, key, label, sd in zip(spec_cols, spec_keys, spec_labels, spec_data):
        ci = openpyxl.utils.column_index_from_string(sc)
        # Merge 3 cols for label
        ws.merge_cells(start_row=3, start_column=ci, end_row=3, end_column=ci+2)
        cell = ws.cell(3, ci, label)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = fills[key]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 18

        # Sub-headers row 4
        for offset, sub in enumerate(['Osnovica', 'PDV', 'Ukupno']):
            c = ws.cell(4, ci+offset, sub)
            c.font = Font(bold=True, size=8, color="FFFFFF")
            c.fill = fills[key]
            c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[4].height = 14

        # Values row 5
        for offset, val in enumerate([sd['osnov'], sd['pdv'], sd['ukupno']]):
            c = ws.cell(5, ci+offset, val)
            c.number_format = num_fmt
            c.font = fonts_col[key]
            c.fill = bg[key]
            c.alignment = Alignment(horizontal='right')
            c.border = border

        # Count badge - merge next to label, row 3
        cnt_c = ws.cell(3, ci+2)  # overwrite last col of merge
        # Put count in row 5 col ci+2 as separate label
        cnt = ws.cell(5, ci+2, sd['ukupno'])  # already set above
        # Add count in row 4 last col
        bc = ws.cell(4, ci+2, f"Ukupno ({sd['count']} kom.)")
        bc.font = Font(bold=True, size=8, color="FFFFFF")
        bc.fill = fills[key]
        bc.alignment = Alignment(horizontal='center')

    ws.row_dimensions[5].height = 16

    # Row 6: empty separator
    ws.row_dimensions[6].height = 8

    # ── TABLE STARTS ROW 7 ────────────────────────────────────────────────────
    HDR_ROW = 7

    # Header style
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="1565C0")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["Broj", "Datum izdavanja", "Datum valute", "Kupac", "PIB",
               "JBKJS", "Događaj", "Osnov (RSD)", "PDV (RSD)", "Ukupno (RSD)",
               "Status", "Napomena", "Korisnik"]
    col_widths = [14, 14, 14, 38, 14, 12, 36, 14, 12, 14, 12, 22, 16]

    ws.row_dimensions[HDR_ROW].height = 30
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=h)
        cell.font    = hdr_font
        cell.fill    = hdr_fill
        cell.alignment = hdr_align
        cell.border  = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Data rows
    alt_fill   = PatternFill("solid", fgColor="EEF3FB")
    norm_fill  = PatternFill("solid", fgColor="FFFFFF")
    green_fill = PatternFill("solid", fgColor="E8F5E9")
    red_fill   = PatternFill("solid", fgColor="FFEBEE")
    num_align  = Alignment(horizontal="right")

    for ri, r in enumerate(rows, HDR_ROW + 1):
        status = (r['status'] or '').upper()
        if 'PLA' in status:
            row_fill  = green_fill
            row_font  = Font(color="1B5E20", size=10)
            num_font  = Font(color="1B5E20", size=10, bold=True)
        elif 'STORNO' in status:
            row_fill  = red_fill
            row_font  = Font(color="CC0000", size=10)
            num_font  = Font(color="CC0000", size=10, bold=True)
        else:
            row_fill  = alt_fill if ri % 2 == 0 else norm_fill
            row_font  = Font(size=10)
            num_font  = Font(size=10, bold=True)

        data = [
            r['broj'], r['datum_izd'], r['datum_val'],
            r['kupac_naziv'], r['kupac_pib'], r['kupac_jbkjs'],
            r['dogadjaj_naziv'],
            float(r['osnov'] or 0), float(r['pdv'] or 0), float(r['ukupno'] or 0),
            r['status'], r['napomena'] or '', r['korisnik_ime'] or ''
        ]
        for ci, val in enumerate(data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill   = row_fill
            cell.border = border
            cell.font   = num_font if isinstance(val, float) else row_font
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
                cell.alignment = num_align
            else:
                cell.alignment = Alignment(vertical="center")

    # Totals row
    tr = HDR_ROW + 1 + len(rows)
    ws.row_dimensions[tr].height = 20
    total_font = Font(bold=True, size=10)
    total_fill = PatternFill("solid", fgColor="D0E4FF")
    ws.cell(tr, 1, "UKUPNO").font = total_font
    ws.cell(tr, 1).fill = total_fill
    for ci in [8, 9, 10]:
        col_letter = get_column_letter(ci)
        cell = ws.cell(tr, ci,
            value=f"=SUM({col_letter}{HDR_ROW+1}:{col_letter}{tr-1})")
        cell.number_format = '#,##0.00'
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = num_align
        cell.border = border

    # Freeze at table header
    ws.freeze_panes = f"A{HDR_ROW+1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import datetime
    fname = f"Evidencija_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


@app.route('/evidencija/export/pdf')
@login_required
def export_pdf():
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos
    import os

    rows = _get_evidencija_rows(request)

    BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
    FONT_DIR  = os.path.join(BASE_DIR, 'static', 'fonts')
    LOGO_PATH = os.path.join(BASE_DIR, 'static', 'img', 'logo.png')

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.set_margins(10, 10, 10)
    pdf.add_page()
    pdf.add_font('R', '', os.path.join(FONT_DIR, 'DejaVuSans.ttf'))
    pdf.add_font('B', '', os.path.join(FONT_DIR, 'DejaVuSans-Bold.ttf'))

    W = 277  # A4 landscape usable width

    # Header
    if os.path.exists(LOGO_PATH):
        pdf.image(LOGO_PATH, x=10, y=6, h=14)
    pdf.set_font('B', '', 13)
    pdf.cell(W, 8, "EVIDENCIJA PREDRAČUNA", align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font('R', '', 8)
    from datetime import datetime
    pdf.cell(W, 5, f"Datum izvoza: {datetime.now().strftime('%d.%m.%Y. %H:%M')}  |  Broj predračuna: {len(rows)}", align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(2)

    # ── SPECIFIKACIJA ─────────────────────────────────────────────────────────
    def sums_for_pdf(status_check):
        matched = [r for r in rows if status_check(str(r['status'] or '').upper())]
        return {'count': len(matched),
                'osnov':  sum(float(r['osnov'] or 0) for r in matched),
                'pdv':    sum(float(r['pdv']   or 0) for r in matched),
                'ukupno': sum(float(r['ukupno'] or 0) for r in matched)}

    specs = [
        ("SVI",        sums_for_pdf(lambda s: True),                     (21,101,192),  (238,243,251)),
        ("AKTIVNI",    sums_for_pdf(lambda s: s not in ('PLAĆENA','STORNO')), (230,81,0),   (255,243,224)),
        ("PLAĆENI",    sums_for_pdf(lambda s: 'PLA' in s),               (46,125,50),   (232,245,233)),
        ("STORNIRANI", sums_for_pdf(lambda s: s == 'STORNO'),            (198,40,40),   (255,235,238)),
    ]

    spec_w = W / 4
    y_spec = pdf.get_y()

    for i, (label, sd, clr, bg_clr) in enumerate(specs):
        x = 10 + i * spec_w
        # Colored header bar
        pdf.set_fill_color(*clr)
        pdf.set_text_color(255,255,255)
        pdf.set_font('B', '', 7.5)
        pdf.rect(x, y_spec, spec_w - 1, 7, 'F')
        pdf.set_xy(x, y_spec)
        pdf.cell(spec_w - 1, 7, f"{label} ({sd['count']} kom.)", align='C')

        # Values background
        pdf.set_fill_color(*bg_clr)
        pdf.rect(x, y_spec + 7, spec_w - 1, 14, 'F')

        # Sub labels
        sub_w = (spec_w - 1) / 3
        pdf.set_font('R', '', 6.5)
        pdf.set_text_color(100,100,100)
        for j, sub in enumerate(["Osnov","PDV","Ukupno"]):
            pdf.set_xy(x + j * sub_w, y_spec + 8)
            pdf.cell(sub_w, 4, sub, align='C')

        # Values
        pdf.set_font('B', '', 7.5)
        pdf.set_text_color(*clr)
        for j, val in enumerate([sd['osnov'], sd['pdv'], sd['ukupno']]):
            pdf.set_xy(x + j * sub_w, y_spec + 12)
            txt = "{:,.2f}".format(val).replace(",","X").replace(".",",").replace("X",".")
            pdf.cell(sub_w, 4, txt, align='C')

    pdf.set_text_color(0, 0, 0)
    pdf.set_y(y_spec + 23)
    pdf.ln(3)

    # Table columns: Broj, Datum izd, Datum val, Kupac, PIB, JBKJS, Događaj, Osnov, PDV, Ukupno, Status
    cw     = [20,  18,  18,  62,  20,  16,  55,  22,  18,  22,  16]
    hdrs   = ["Broj","Datum izd.","Datum val.","Kupac","PIB","JBKJS","Događaj","Osnov","PDV","Ukupno","Status"]
    aligns = ['C','C','C','L','C','C','L','R','R','R','C']

    # Table header
    pdf.set_fill_color(21, 101, 192)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('B', '', 7.5)
    pdf.set_draw_color(180, 180, 180)
    pdf.set_line_width(0.2)
    for h, w, a in zip(hdrs, cw, aligns):
        pdf.cell(w, 7, h, border=1, align=a, fill=True, new_x=XPos.RIGHT, new_y=YPos.TOP)
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font('R', '', 7.5)

    def fmt(x):
        try: return "{:,.2f}".format(float(x or 0)).replace(",","X").replace(".",",").replace("X",".")
        except: return ""

    alt = False
    osnov_sum = pdv_sum = ukupno_sum = 0.0

    for r in rows:
        alt = not alt
        if alt:
            pdf.set_fill_color(238, 243, 251)
        else:
            pdf.set_fill_color(255, 255, 255)

        vals = [
            r['broj'] or '',
            r['datum_izd'] or '',
            r['datum_val'] or '',
            r['kupac_naziv'] or '',
            str(r['kupac_pib'] or ''),
            str(r['kupac_jbkjs'] or ''),
            r['dogadjaj_naziv'] or '',
            fmt(r['osnov']),
            fmt(r['pdv']),
            fmt(r['ukupno']),
            r['status'] or '',
        ]

        # Measure row height from kupac and dogadjaj cols
        import textwrap
        kupac_lines = len(textwrap.wrap(vals[3], 28)) or 1
        dog_lines   = len(textwrap.wrap(vals[6], 24)) or 1
        row_h = max(kupac_lines, dog_lines) * 4.5
        row_h = max(row_h, 6)

        y0 = pdf.get_y()
        if y0 + row_h > pdf.page_break_trigger:
            pdf.add_page()
            # Redraw header
            pdf.set_fill_color(21, 101, 192)
            pdf.set_text_color(255,255,255)
            pdf.set_font('B', '', 7.5)
            for h, w, a in zip(hdrs, cw, aligns):
                pdf.cell(w, 7, h, border=1, align=a, fill=True, new_x=XPos.RIGHT, new_y=YPos.TOP)
            pdf.ln()
            pdf.set_text_color(0,0,0)
            pdf.set_font('R', '', 7.5)
            y0 = pdf.get_y()

        x0 = 10
        for val, w, a in zip(vals, cw, aligns):
            if a == 'L' and len(val) > 15:
                pdf.rect(x0, y0, w, row_h, 'DF')
                pdf.set_xy(x0 + 0.5, y0 + 0.5)
                pdf.multi_cell(w - 1, 4.5, val, border=0, align='L')
            else:
                pdf.set_fill_color(*([238,243,251] if alt else [255,255,255]))
                pdf.rect(x0, y0, w, row_h, 'DF')
                pdf.set_xy(x0, y0)
                pdf.cell(w, row_h, val, border=0, align=a)
            x0 += w
        pdf.set_xy(10, y0 + row_h)

        osnov_sum   += float(r['osnov'] or 0)
        pdv_sum     += float(r['pdv'] or 0)
        ukupno_sum  += float(r['ukupno'] or 0)

    # Totals row
    pdf.set_fill_color(208, 228, 255)
    pdf.set_font('B', '', 8)
    total_vals = ["UKUPNO","","","","","","",fmt(osnov_sum),fmt(pdv_sum),fmt(ukupno_sum),""]
    for val, w, a in zip(total_vals, cw, aligns):
        pdf.cell(w, 7, val, border=1, align=a, fill=True, new_x=XPos.RIGHT, new_y=YPos.TOP)
    pdf.ln()

    buf = io.BytesIO(bytes(pdf.output()))
    fname = f"Evidencija_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=fname)

@app.route('/predracun/uredi/<int:pid>')
@login_required
def uredi_predracun(pid):
    p = q("SELECT * FROM evidencija WHERE id=?", (pid,), one=True)
    if not p: abort(404)
    dogadjaji = q("SELECT * FROM dogadjaji WHERE status=\'AKTIVNA\' ORDER BY id DESC")
    cenovnik  = q("SELECT * FROM cenovnik WHERE status=\'AKTIVNA\' ORDER BY naziv")
    return render_template('uredi_predracun.html',
        p=dict(p), dogadjaji=dogadjaji, cenovnik=cenovnik)

@app.route('/predracun/azuriraj/<int:pid>', methods=['POST'])
@login_required
def azuriraj_predracun(pid):
    data = request.json
    stavke_json = json.dumps(data.get('stavke', []), ensure_ascii=False)
    osnov = 0.0; pdv_t = 0.0
    for st in data.get('stavke', []):
        try:
            c=float(st['cena']); k=float(st['kolicina']); p2=float(st['pdv'])
            b=c*k; osnov+=b; pdv_t+=b*p2/100
        except: pass
    ukupno = osnov + pdv_t
    dog_naziv = ''
    if data.get('dogadjaj_id'):
        dog = q("SELECT * FROM dogadjaji WHERE id=?", (data['dogadjaj_id'],), one=True)
        if dog: dog_naziv = dog['naziv']
    specifikacija = data.get('specifikacija','')
    q("""UPDATE evidencija SET
        broj=?, datum_izd=?, datum_val=?,
        kupac_naziv=?, kupac_pib=?, kupac_mb=?, kupac_adresa=?, kupac_mesto=?, kupac_jbkjs=?,
        kupac_extra=?, dogadjaj_id=?, dogadjaj_naziv=?,
        osnov=?, pdv=?, ukupno=?, napomena_predracun=?, specifikacija=?, stavke=?
        WHERE id=?""",
        (data['broj'], data['datum_izd'], data['datum_val'],
         data['kupac_naziv'], data.get('kupac_pib',''), data.get('kupac_mb',''),
         data.get('kupac_adresa',''), data.get('kupac_mesto',''), data.get('kupac_jbkjs',''),
         data.get('kupac_extra',''),
         data.get('dogadjaj_id') or None, dog_naziv,
         osnov, pdv_t, ukupno, data.get('napomena',''), specifikacija, stavke_json, pid), commit=True)
    return jsonify({'ok': True, 'id': pid})

if __name__ == '__main__':
    # Auto-initialize DB if missing
    if not os.path.exists(DB_PATH):
        print("Inicijalizacija baze podataka...")
        from init_db import init_db
        init_db()
    print("=" * 45)
    print("  CMN Predracun Manager")
    print("  Otvorite: http://localhost:5000")
    print("  Prijava:  admin / admin123")
    print("  Ctrl+C za zaustavljanje")
    print("=" * 45)
    try:
        from waitress import serve
        serve(app, host='127.0.0.1', port=5000, threads=4)
    except ImportError:
        app.run(debug=False, host='127.0.0.1', port=5000)


@app.route('/admin/import', methods=['GET','POST'])
@login_required
@admin_required
def import_data():
    if request.method == 'GET':
        return render_template('import.html')
    import pandas as pd, json as json2
    results = []
    try:
        f2 = request.files.get('file')
        if not f2:
            return jsonify({'error': 'Nema fajla'})
        xl = pd.ExcelFile(f2)

        # Obrisi sve pre uvoza
        q("DELETE FROM kupci", commit=True)
        q("DELETE FROM cenovnik", commit=True)
        q("DELETE FROM dogadjaji", commit=True)
        q("DELETE FROM evidencija", commit=True)

        # Kupci - batch insert za brzinu
        if 'Kupci' in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name='Kupci')
            df.columns = [col.strip().upper() for col in df.columns]
            rows_to_insert = []
            for _, row in df.iterrows():
                naziv = str(row.get('NAZIV', '')).strip()
                if not naziv or naziv == 'nan': continue
                pib = str(row.get('PIB', '')).replace('.0','').strip()
                mb  = str(row.get('MB', '')).replace('.0','').strip()
                jbkjs = str(row.get('JBKJS', '')).split('.')[0].strip()
                adresa = str(row.get('ADRESA','')).strip()
                mesto = str(row.get('MESTO','')).strip()
                for v in ('nan','None','NaN'):
                    if pib==v: pib=''
                    if mb==v: mb=''
                    if jbkjs==v: jbkjs=''
                    if adresa==v: adresa=''
                    if mesto==v: mesto=''
                rows_to_insert.append((naziv, pib, mb, adresa, mesto, jbkjs))
            # Chunked insert - 200 rows at a time to avoid memory issues
            import os as _os2
            _db_url2 = _os2.environ.get('DATABASE_URL','')
            chunk_size = 200
            total = 0
            for i in range(0, len(rows_to_insert), chunk_size):
                chunk = rows_to_insert[i:i+chunk_size]
                if _db_url2:
                    import psycopg2, psycopg2.extras
                    conn2 = psycopg2.connect(_db_url2)
                    cur2 = conn2.cursor()
                    psycopg2.extras.execute_batch(cur2,
                        "INSERT INTO kupci (naziv, pib, mb, adresa, mesto, jbkjs) VALUES (%s,%s,%s,%s,%s,%s)",
                        chunk, page_size=200)
                    conn2.commit(); conn2.close()
                else:
                    import sqlite3 as _sq
                    db_path2 = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance', 'cmn.db')
                    conn2 = _sq.connect(db_path2)
                    conn2.executemany("INSERT INTO kupci (naziv, pib, mb, adresa, mesto, jbkjs) VALUES (?,?,?,?,?,?)", chunk)
                    conn2.commit(); conn2.close()
                total += len(chunk)
            results.append(f"Kupci: {total} uvezeno")

        # Cenovnik
        if 'Cenovnik' in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name='Cenovnik')
            df.columns = [col.strip().upper() for col in df.columns]
            count = 0
            for _, row in df.iterrows():
                naziv = str(row.get('NAZIV', '')).strip()
                if not naziv or naziv == 'nan': continue
                try:
                    cena = float(str(row.get('CENA',0)).replace(',','.'))
                    pdv2 = int(float(str(row.get('PDV',20))))
                    status2 = str(row.get('STATUS','AKTIVNA')).strip()
                    if status2 == 'nan': status2 = 'AKTIVNA'
                    q("INSERT INTO cenovnik (naziv, cena, pdv, status) VALUES (?,?,?,?)",
                      (naziv, cena, pdv2, status2), commit=True)
                    count += 1
                except: pass
            results.append(f"Cenovnik: {count} uvezeno")

        # Dogadjaji (sheet Opis, kolone: ID_DOG, DOGADJAJ, DATUM, STATUS)
        if 'Opis' in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name='Opis')
            df.columns = [col.strip().upper() for col in df.columns]
            count = 0
            for _, row in df.iterrows():
                naziv = str(row.get('DOGADJAJ', '')).strip()
                if not naziv or naziv == 'nan': continue
                datum = str(row.get('DATUM', '')).strip()
                if datum == 'nan': datum = ''
                status3 = str(row.get('STATUS', 'AKTIVNA')).strip()
                if status3 == 'nan': status3 = 'AKTIVNA'
                try:
                    q("INSERT INTO dogadjaji (naziv, datum, status) VALUES (?,?,?)",
                      (naziv, datum, status3), commit=True)
                    count += 1
                except: pass
            results.append(f"Dogadjaji: {count} uvezeno")

        # Evidencija
        if 'Evidencija' in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name='Evidencija')
            df.columns = [col.strip().upper() for col in df.columns]
            count = 0
            for _, row in df.iterrows():
                broj = str(row.get('BROJ', '')).strip()
                if not broj or broj == 'nan': continue
                def sv2(col, default=''):
                    val = row.get(col, default)
                    try:
                        if pd.isna(val): return default
                    except: pass
                    s = str(val).strip()
                    return default if s in ('nan','None','NaN') else s
                def fv2(col):
                    try: return float(row.get(col, 0))
                    except: return 0.0
                stavke_tekst = sv2('STAVKE')
                # Parse "naziv x{kol} @ {cena}" format
                stavke_list = []
                import re
                m = re.match(r'^(.+?)\s+x(\d+)\s+@\s+([\d.]+)$', stavke_tekst)
                if m:
                    stavke_list = [{"naziv": m.group(1).strip(), "kolicina": int(m.group(2)), "cena": float(m.group(3)), "pdv": 20}]
                else:
                    stavke_list = [{"naziv": stavke_tekst, "kolicina": 1, "cena": fv2('OSNOV'), "pdv": 20}]
                stavke_json = json2.dumps(stavke_list, ensure_ascii=False)
                try:
                    q("INSERT INTO evidencija (broj, datum_izd, datum_val, kupac_naziv, kupac_pib, kupac_jbkjs, dogadjaj_id, dogadjaj_naziv, osnov, pdv, ukupno, status, napomena, stavke, korisnik_ime) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                      (broj, sv2('DATUM_IZD'), sv2('DATUM_VAL'), sv2('KUPAC_NAZIV'),
                       sv2('KUPAC_PIB').replace('.0',''),
                       str(int(float(sv2('KUPAC_JBKJS','0') or '0'))) if sv2('KUPAC_JBKJS') else '',
                       int(float(sv2('DOGADJAJ_ID','0') or '0')) if sv2('DOGADJAJ_ID') else None,
                       sv2('DOGADJAJ_NAZIV'),
                       fv2('OSNOV'), fv2('PDV'), fv2('UKUPNO'),
                       sv2('STATUS','ČEKA'), sv2('NAPOMENA'),
                       stavke_json, 'Import'), commit=True)
                    count += 1
                except Exception as ex:
                    pass
            results.append(f"Evidencija: {count} uvezeno")

        return jsonify({'ok': True, 'results': results})
    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()})
